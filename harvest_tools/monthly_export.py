"""Export HEV004 time entries from Harvest to XLSX and email to client monthly."""

import os
from datetime import date, datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Font

from .client import get_client
from .email import send_email
from .telegram import send_message

PROJECT_ID = 47325879  # HEV004 — System Admin Services
TASK_ID = 26287739  # Tracks 4 Africa
CALENDAR_NAME = "T4A"


def _month_range(ref_date: date) -> tuple[date, date]:
    """Return first and last day of the month containing ref_date."""
    first = ref_date.replace(day=1)
    if first.month == 12:
        last = first.replace(year=first.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        last = first.replace(month=first.month + 1, day=1) - timedelta(days=1)
    return first, last


def _fetch_entries(start: date, end: date, task_id: int | None = None) -> list[dict]:
    """Fetch time entries for HEV004 in the date range, optionally filtered by task."""
    client = get_client()
    entries = []
    page = 1
    while True:
        params = {
            "project_id": PROJECT_ID,
            "from": str(start),
            "to": str(end),
            "per_page": 100,
            "page": page,
        }
        if task_id is not None:
            params["task_id"] = task_id
        resp = client.get("/time_entries", params=params)
        resp.raise_for_status()
        data = resp.json()
        entries.extend(data["time_entries"])
        if page >= data["total_pages"]:
            break
        page += 1
    return entries


def _hours_to_time_str(hours: float) -> str:
    """Convert decimal hours to HH:MM string."""
    total_minutes = round(hours * 60)
    h, m = divmod(total_minutes, 60)
    return f"{h}:{m:02d}"


def _parse_start_time(entry: dict) -> datetime | None:
    """Get start time from started_time (timer) or created_at (fallback)."""
    spent = date.fromisoformat(entry["spent_date"])
    if entry.get("started_time"):
        t = datetime.strptime(entry["started_time"], "%I:%M%p").time()
        return datetime.combine(spent, t)
    if entry.get("created_at"):
        sast = timezone(timedelta(hours=2))
        utc_dt = datetime.fromisoformat(entry["created_at"].replace("Z", "+00:00"))
        return utc_dt.astimezone(sast).replace(tzinfo=None)
    return None


def _fmt_dt(dt: datetime | None) -> str | None:
    """Format a datetime as 'DD/MM/YYYY HH:MM', or None."""
    if dt is None:
        return None
    return dt.strftime("%d/%m/%Y %H:%M")


def _build_t4a_workbook(entries: list[dict]) -> Workbook:
    """Build an XLSX workbook matching the T4A report template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "Calendar Name",
        "Event Title",
        "Start Date/Time",
        "End Date/Time",
        "Duration (Hours, per 15 minutes commenced)",
        "Event Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    total_seconds = 0
    for e in entries:
        hours = e["hours"]
        notes = (e.get("notes") or "").strip()
        start = _parse_start_time(e)
        end = start + timedelta(hours=hours) if start else None

        ws.append([
            CALENDAR_NAME,
            notes,
            _fmt_dt(start),
            _fmt_dt(end),
            _hours_to_time_str(hours),
            "",
        ])
        total_seconds += round(hours * 3600)

    total_row = ws.max_row + 1
    total_h, remainder = divmod(total_seconds, 3600)
    total_m = remainder // 60
    ws.cell(row=total_row, column=5, value=f"{total_h}:{total_m:02d}")

    return wb


def _build_full_workbook(entries: list[dict]) -> Workbook:
    """Build an XLSX workbook with all HEV004 entries."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "Date",
        "Task",
        "Description",
        "Start Time",
        "End Time",
        "Duration",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    total_seconds = 0
    for e in entries:
        hours = e["hours"]
        task_name = e.get("task", {}).get("name", "")
        notes = (e.get("notes") or "").strip()
        start = _parse_start_time(e)
        end = start + timedelta(hours=hours) if start else None

        ws.append([
            e["spent_date"],
            task_name,
            notes,
            _fmt_dt(start),
            _fmt_dt(end),
            _hours_to_time_str(hours),
        ])
        total_seconds += round(hours * 3600)

    total_row = ws.max_row + 1
    total_h, remainder = divmod(total_seconds, 3600)
    total_m = remainder // 60
    ws.cell(row=total_row, column=6, value=f"{total_h}:{total_m:02d}")

    return wb


def main():
    today = date.today()
    start, end = _month_range(today)
    month_label = start.strftime("%B %Y")
    client_email = os.environ["CLIENT_EMAIL_RECIPIENT"]

    print(f"Fetching HEV004 entries for {month_label}...")

    # Full project export
    all_entries = _fetch_entries(start, end)
    if all_entries:
        full_filename = f"hev004-{today.strftime('%Y-%m')}.xlsx"
        _build_full_workbook(all_entries).save(full_filename)
        print(f"Exported {len(all_entries)} entries to {full_filename}")

        send_email(
            subject=f"HEV004 Monthly Report — {month_label}",
            body=f"Please find attached the full HEV004 time report for {month_label}.",
            recipient=client_email,
            attachments=[full_filename],
        )
        print(f"Full report emailed to client.")

    # T4A-specific export
    t4a_entries = _fetch_entries(start, end, task_id=TASK_ID)
    if not t4a_entries:
        msg = f"T4A Export — {month_label}: No Tracks 4 Africa entries found for this month."
        print(msg)
        send_message(msg)
        return

    t4a_filename = f"tracks4africa-{today.strftime('%Y-%m')}.xlsx"
    _build_t4a_workbook(t4a_entries).save(t4a_filename)
    print(f"Exported {len(t4a_entries)} T4A entries to {t4a_filename}")

    send_email(
        subject=f"Tracks 4 Africa Report — {month_label}",
        body=f"Please find attached the Tracks 4 Africa time report for {month_label}.",
        recipient=client_email,
        attachments=[t4a_filename],
    )
    print(f"T4A report emailed to client.")


if __name__ == "__main__":
    main()
